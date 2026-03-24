import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load data
rv = pd.read_parquet('rv_daily.parquet')
rv['timestamp'] = pd.to_datetime(rv['timestamp']).dt.date

dm = pd.read_excel('strategy_returns_DM_per_trade_both_max_100.xlsx', sheet_name='returns')
dm['Date'] = pd.to_datetime(dm['Date']).dt.date

wc = pd.read_excel('strategy_returns_90_0_both_itm.xlsx', sheet_name='returns')
wc['Date'] = pd.to_datetime(wc['Date']).dt.date

orion = pd.read_excel('strategy_returns_orion_index_kd_60_40_sl10_max90_min20.xlsx', sheet_name='returns')
orion['Date'] = pd.to_datetime(orion['Date']).dt.date

# Merge all on date
merged = rv.rename(columns={'timestamp': 'Date'}).copy()
merged = merged.merge(
    dm[['Date', 'Net_Daily_PnL_PerCent', 'Net_Equity_Curve', 'Net_PnL', 'Trades']].rename(
        columns={'Net_Daily_PnL_PerCent': 'DM_Return_%', 'Net_Equity_Curve': 'DM_Equity_Curve',
                 'Net_PnL': 'DM_Net_PnL', 'Trades': 'DM_Trades'}),
    on='Date', how='left'
)
merged = merged.merge(
    wc[['Date', 'Net_Daily_PnL_PerCent', 'Net_Equity_Curve', 'Net_PnL', 'Trades']].rename(
        columns={'Net_Daily_PnL_PerCent': 'WC_Return_%', 'Net_Equity_Curve': 'WC_Equity_Curve',
                 'Net_PnL': 'WC_Net_PnL', 'Trades': 'WC_Trades'}),
    on='Date', how='left'
)
merged = merged.merge(
    orion[['Date', 'Net_Daily_PnL_PerCent', 'Net_Equity_Curve', 'Net_PnL', 'Trades']].rename(
        columns={'Net_Daily_PnL_PerCent': 'Orion_Return_%', 'Net_Equity_Curve': 'Orion_Equity_Curve',
                 'Net_PnL': 'Orion_Net_PnL', 'Trades': 'Orion_Trades'}),
    on='Date', how='left'
)

# Build Excel
wb = Workbook()
ws = wb.active
ws.title = 'Daily Overview'

# Styles
header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
section_fill_rv = PatternFill('solid', fgColor='4472C4')
section_fill_iv = PatternFill('solid', fgColor='7030A0')
section_fill_dm = PatternFill('solid', fgColor='548235')
section_fill_wc = PatternFill('solid', fgColor='BF8F00')
section_fill_orion = PatternFill('solid', fgColor='C00000')
data_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Section headers (row 1)
sections = [
    ('A1', 'Nifty OHLC', 5, header_fill),
    ('F1', 'RV Features', 5, section_fill_rv),
    ('K1', 'IV Features', 3, section_fill_iv),
    ('N1', 'DM Strategy', 4, section_fill_dm),
    ('R1', 'WC Strategy', 4, section_fill_wc),
    ('V1', 'Orion Strategy', 4, section_fill_orion),
]
for cell, title, span, fill in sections:
    ws[cell] = title
    ws[cell].font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    ws[cell].fill = fill
    ws[cell].alignment = Alignment(horizontal='center')
    col = ws[cell].column
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
    for c in range(col, col + span):
        ws.cell(row=1, column=c).fill = fill

# Column headers (row 2)
headers = [
    'Date', 'Open', 'High', 'Low', 'Close',
    'RV_today', 'RV_3d_avg', 'RV_ratio', 'RV_7d_avg', 'RV_7d_ratio',
    'IV_7d', 'IV_change_1d', 'VRP_today',
    'DM_Net_Return_%', 'DM_Net_Equity_Curve', 'DM_Net_PnL', 'DM_Trades',
    'WC_Net_Return_%', 'WC_Net_Equity_Curve', 'WC_Net_PnL', 'WC_Trades',
    'Orion_Net_Return_%', 'Orion_Net_Equity_Curve', 'Orion_Net_PnL', 'Orion_Trades',
]

# Map column index to section fill
section_colors = {}
for i in range(5):
    section_colors[i] = header_fill
for i in range(5, 10):
    section_colors[i] = section_fill_rv
for i in range(10, 13):
    section_colors[i] = section_fill_iv
for i in range(13, 17):
    section_colors[i] = section_fill_dm
for i in range(17, 21):
    section_colors[i] = section_fill_wc
for i in range(21, 25):
    section_colors[i] = section_fill_orion

for i, h in enumerate(headers):
    cell = ws.cell(row=2, column=i+1, value=h)
    cell.font = header_font
    cell.fill = section_colors[i]
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Data columns mapping
cols = ['Date', 'open', 'high', 'low', 'close',
        'RV_today', 'RV_3d_avg', 'RV_ratio', 'RV_7d_avg', 'RV_7d_ratio',
        'IV_7d', 'IV_change_1d', 'VRP_today',
        'DM_Return_%', 'DM_Equity_Curve', 'DM_Net_PnL', 'DM_Trades',
        'WC_Return_%', 'WC_Equity_Curve', 'WC_Net_PnL', 'WC_Trades',
        'Orion_Return_%', 'Orion_Equity_Curve', 'Orion_Net_PnL', 'Orion_Trades']

even_fill = PatternFill('solid', fgColor='F2F2F2')

for row_idx, (_, row) in enumerate(merged.iterrows()):
    r = row_idx + 3
    for col_idx, col_name in enumerate(cols):
        val = row.get(col_name)
        if pd.isna(val):
            val = None
        cell = ws.cell(row=r, column=col_idx+1, value=val)
        cell.font = data_font
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = even_fill

# Number formatting
n_cols = len(headers)
for r in range(3, len(merged) + 3):
    ws.cell(row=r, column=1).number_format = 'YYYY-MM-DD'
    for c in [2, 3, 4, 5]:
        ws.cell(row=r, column=c).number_format = '#,##0.00'
    # RV columns: RV_today, RV_3d_avg, RV_7d_avg
    for c in [6, 7, 9]:
        ws.cell(row=r, column=c).number_format = '0.00'
    # RV ratios
    for c in [8, 10]:
        ws.cell(row=r, column=c).number_format = '0.000'
    # IV_7d, IV_change_1d, VRP_today
    for c in [11, 12, 13]:
        ws.cell(row=r, column=c).number_format = '0.00'
    # Strategy return %
    for c in [14, 18, 22]:
        ws.cell(row=r, column=c).number_format = '0.00%'
    # Strategy equity curve
    for c in [15, 19, 23]:
        ws.cell(row=r, column=c).number_format = '0.00%'
    # Strategy net PnL
    for c in [16, 20, 24]:
        ws.cell(row=r, column=c).number_format = '#,##0.00'
    # Strategy trades
    for c in [17, 21, 25]:
        ws.cell(row=r, column=c).number_format = '0'

# Column widths
widths = {}
for i in range(1, n_cols + 1):
    widths[i] = 14
widths[1] = 12  # Date
for c in [17, 21, 25]:  # Trades columns
    widths[c] = 8

for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = 'B3'
ws.auto_filter.ref = f'A2:{get_column_letter(n_cols)}{len(merged)+2}'

out = 'daily_overview_all_strategies.xlsx'
wb.save(out)
print(f'Saved: {out} | Rows: {len(merged)}, Columns: {len(headers)}')
